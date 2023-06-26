from openpyxl import load_workbook
import pandas as pd
from glob import glob
import logging.config
import shutil
import re
import os
import logging.config

from save_xlsx_sheets_as_pdf import save_xlsx_sheets_as_pdf
from spur_xlsx_html_formatting import *


def xlsx_write_up_extract(
    xlsx_list,
    # xlsx_destination,
    save_xlsx_sheet,
    save_xml,
    data_dir,
    job_blob_dir,
    job_clob_dir,
    position_blob_dir,
    position_clob_dir,
    skg_name,
):
    xlsx_list = [x for x in xlsx_list if not re.search(r"~\$", x)]
    final_write_up_list = []
    for xlsx_file in xlsx_list:
        print(xlsx_file)

        if save_xml == True:
            save_excel_as_xml(xlsx_file=xlsx_file, data_dir=data_dir)

        wb = load_workbook(xlsx_file)
        if skg_name == "SKG10":
            do_deliver_re = re.compile("DISPLAY|GTI|All")
        elif skg_name == "SKG016":
            do_deliver_re = re.compile("^\d|SUMMARY|DISPLAY", flags=re.I)
        else:
            do_deliver_re = re.compile(
                "DISPLAY|TI&R|^Sheet|^Summary|^SUMMARY|^Approval|"
                "^Sum of Position|^JCP|^Competency|^Corporate Investment|^Treasury|^SPUR->|"
                "^Listing from HR|Clustering",
                flags=re.I,
            )

        write_up_list = []
        ws_name_list = []
        for ws_name in wb.sheetnames:
            if not do_deliver_re.search(ws_name):
                # print(ws_name)
                ws = wb[ws_name]
                if ws.sheet_state != "visible":
                    continue

                df_list = []
                # Iterating rows for getting the values of each row
                for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=30):
                    df_list.append([cell.value for cell in row])

                initial_df = (
                    pd.DataFrame(df_list).rename(columns=lambda x: str(x).strip()).apply(lambda x: x.str.strip())
                )
                try:
                    try:
                        columns_index = initial_df.index[initial_df.iloc[:, 0] == "Role Purpose"][0]
                    except:
                        try:
                            columns_index = initial_df.index[initial_df.iloc[:, 1] == "Role Purpose"][0]
                        except:
                            try:
                                columns_index = initial_df.index[initial_df.iloc[:, 2] == "Role Purpose"][0]
                            except:
                                columns_index = initial_df.index[initial_df.iloc[:, 3] == "Role Purpose"][0]
                except:
                    print(f"Sheet {ws_name} has no write-up")
                    continue

                df = (
                    pd.DataFrame(df_list)
                    .T.set_index(columns_index)
                    .T.reset_index(drop=True)
                    .pipe(lambda x: x.loc[:, ~x.columns.duplicated()])
                )
                # df.columns = [" ".join(str(x).strip().split()) for x in df.columns]
                if skg_name != "SKG001":
                    ur_code_name = [x for x in df_list[0] if x is not None][0]
                else:
                    ur_id_column = [col for col in df.columns if re.search("Unique Role ID|UR ID", str(col))][0]
                    ur_code_name = (
                        df[ur_id_column].iloc[columns_index].strip()
                        + "|"
                        + df["Unique Role"].iloc[columns_index].strip()
                    ).strip()

                ur_code = (
                    re.search(".+?(?=\|)", ur_code_name, flags=re.S)
                    .group()
                    .strip()
                    .replace("SKG", "0")
                    .replace("UR ID: ", "")
                )
                ur_name = re.search("[^\|]+$", ur_code_name).group().strip()
                ur_name = re.sub("-$|\[.+", "", ur_name, flags=re.S).strip()
                if ur_name.isupper() == True:
                    ur_name = ur_name.title()
                source = re.search(r"[^\\]+$", xlsx_file).group().strip()
                columns = [x for x in df.columns if x != None]
                # print(columns)
                for ur in re.findall("\d+\s*\-\s*\d+", ur_code):
                    df2 = (
                        df.loc[columns_index:, columns]
                        .reset_index(drop=True)
                        .rename(columns=lambda x: str(x).strip())
                        .pipe(lambda x: x.loc[:, ~x.columns.duplicated()])
                        .pipe(lambda x: x.assign(UR_CODE=ur))
                        .pipe(lambda x: x.fillna("(NA)"))
                        .pipe(lambda x: x.groupby("UR_CODE").agg("\n\n".join).reset_index(drop=False))
                        .pipe(lambda x: x.assign(UR_NAME=ur_name, SOURCE=source))
                        .pipe(lambda x: x.replace("\n\(NA\)", "", regex=True))
                        .pipe(
                            lambda x: x.rename(
                                {
                                    "Challenges *": "CHALLENGES",
                                    "Challenges": "CHALLENGES",
                                    "Role Purpose": "ROLEPURPOSE",
                                    "Main Accountabilities": "ACCOUNTABILITIES",
                                    "Accountabilities": "ACCOUNTABILITIES",
                                    "Accountabilties": "ACCOUNTABILITIES",
                                    "Experience": "EXPERIENCE",
                                    "Key Performance Indicator": "KPI",
                                    4: "EXPERIENCE",
                                    "4": "EXPERIENCE",
                                },
                                axis=1,
                            )
                        )
                        .pipe(
                            lambda x: x[
                                [
                                    "UR_CODE",
                                    "UR_NAME",
                                    "ROLEPURPOSE",
                                    "ACCOUNTABILITIES",
                                    "CHALLENGES",
                                    "EXPERIENCE",
                                    "KPI",
                                    "SOURCE",
                                ]
                            ]
                        )
                        .apply(lambda x: x.str.strip())
                        .replace("\(NA\)", "", regex=True)
                    )
                    df2.columns.name = None
                    df2 = df2.replace("_x000B_", "\n", regex=True)
                    bold_list = get_bold_list(skg_name, xlsx_file)
                    # bold_list = []
                    apply_col = [col for col in df2.columns if not re.search("UR_CODE|UR_NAME|SOURCE", col)]
                    for col in apply_col:
                        df2[col] = df2[col].apply(lambda x: apply_html_format(x, bold_list))
                    write_up_list.append(df2)
                    ws_name_list.append([ws_name, ur])
            #             print(df)
            else:
                continue

        write_up_df = pd.concat(write_up_list).reset_index(drop=True)
        final_write_up_list.append(write_up_df)

        # Save
        if save_xlsx_sheet == True:
            save_xlsx_sheets_as_pdf(
                xlsx_file,
                job_blob_dir,
                job_clob_dir,
                position_blob_dir,
                position_clob_dir,
                skg_name,
                ws_name_list,
            )
        if len(os.listdir(job_blob_dir)) != 0:
            for f in os.listdir(job_blob_dir):
                shutil.copy2(job_blob_dir + "\\" + f, position_blob_dir)

    final_write_up_df = pd.concat(final_write_up_list).reset_index(drop=True)

    # bold_list = get_bold_list(skg_name)

    # apply_col = [col for col in final_write_up_df.columns if not re.search("UR_CODE|UR_NAME|SOURCE", col)]
    # for col in apply_col:
    #     final_write_up_df[col] = final_write_up_df[col].apply(lambda x: apply_html_format(x, bold_list))

    for i in range(len(final_write_up_df)):
        # DESCRIPTION
        with open(
            job_clob_dir + "\\" + final_write_up_df.iloc[i]["UR_CODE"] + "_DESCRIPTION.txt",
            "w",
            encoding="utf-8",
        ) as f:
            f.write(
                "<p>"
                + final_write_up_df.iloc[i]["ROLEPURPOSE"].strip()
                + "</p>"
                + "\n\n\n"
                + final_write_up_df.iloc[i]["ACCOUNTABILITIES"].strip()
            )
        # RESPONSIBILITY
        with open(
            job_clob_dir + "\\" + final_write_up_df.iloc[i]["UR_CODE"] + "_RESPONSIBILITY.txt",
            "w",
            encoding="utf-8",
        ) as f:
            f.write(final_write_up_df.iloc[i]["KPI"].strip())
        # QUALIFICATION
        with open(
            job_clob_dir + "\\" + final_write_up_df.iloc[i]["UR_CODE"] + "_QUALIFICATION.txt",
            "w",
            encoding="utf-8",
        ) as f:
            f.write(final_write_up_df.iloc[i]["EXPERIENCE"].strip())

    if (save_xlsx_sheet == True) & (len(os.listdir(job_clob_dir)) != 0):
        for f in os.listdir(job_clob_dir):
            shutil.copy2(job_clob_dir + "\\" + f, position_clob_dir)

    return final_write_up_df
    # final_write_up_df.to_excel(xlsx_destination, index=False)