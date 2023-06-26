import pandas as pd
from pandas import ExcelWriter
import numpy as np
import re
import itertools
import glob
import os
import sys
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import openpyxl
import string
from openpyxl import Workbook, load_workbook
import logging.config

log = logging.getLogger(__name__)


def trim_all_columns(df):
    """
    Trim whitespace from ends of each value across all series in dataframe
    """
    trim_strings = lambda x: x.strip() if isinstance(x, str) else x
    return df.applymap(trim_strings)


class data_processor:
    """
    docstring
    """

    def __init__(
        self,
        main_dir,
        position_master_data_file_path,
        tc_in_simplified_template,
        simplified_template_file_path,
        tc_raw_path,
        cocode_map_path,
        skg_name,
    ):
        self.main_dir = main_dir
        self.position_master_data_file_path = position_master_data_file_path
        self.tc_in_simplified_template = tc_in_simplified_template
        self.simplified_template_file_path = simplified_template_file_path
        self.tc_raw_path = tc_raw_path
        self.cocode_map_path = cocode_map_path
        self.skg_name = skg_name
        self.UR_ID_list = []

        # get SPUR ID list
        # data_dict = pd.read_excel(self.position_master_data_file_path, sheet_name=None)
        # self.UR_ID_list.extend(list(data_dict.keys()))

    def simplified_template_data(self):
        """
        docstring
        """
        data_path_list = glob.glob(self.simplified_template_file_path)
        data_path_list = [x for x in data_path_list if not re.search(r"~\$", x)]

        exp_list = []
        degree_list = []
        membership_list = []
        awards_list = []
        license_list = []
        for data_path in data_path_list:
            data_dict = pd.read_excel(
                data_path,
                sheet_name=None,
                # converters={"Min Years": int, "Max Years": int, "Importance": int},
            )
            for key in data_dict.keys():
                if "lookup" in key.lower() or "competency" in key.lower():
                    continue

                if "experience" in key.lower() or "experiencerequired" in key.lower():
                    exp_list.append(data_dict[key].dropna(subset=["Importance"]))

                elif "degree" in key.lower():
                    degree_list.append(data_dict[key].dropna(subset=["Importance"]))

                elif "membership" in key.lower():
                    membership_list.append(data_dict[key].dropna(subset=["Importance"]))

                elif "awards" in key.lower():
                    awards_list.append(data_dict[key].dropna(subset=["Importance"]))

                elif "license" in key.lower():
                    license_list.append(data_dict[key].dropna(subset=["Importance"]))

                else:
                    continue

        exp_df = pd.concat(exp_list).reset_index(drop=True)
        exp_df = trim_all_columns(exp_df)
        if not any(["JG" in str(x) for x in exp_df.columns]):
            exp_df["JG"] = np.nan

        degree_df = pd.concat(degree_list).reset_index(drop=True)
        degree_df = trim_all_columns(degree_df)
        if not any(["JG" in str(x) for x in degree_df.columns]):
            degree_df["JG"] = np.nan

        membership_df = pd.concat(membership_list).reset_index(drop=True)
        membership_df = trim_all_columns(membership_df)
        membership_df_obj = membership_df.select_dtypes(["object"])
        membership_df[membership_df_obj.columns] = membership_df_obj.apply(lambda x: x.str.strip())
        membership_column = [
            x
            for x in membership_df.columns
            if re.search(
                "Membership - Affiliation or Professional Body|Bodies membership",
                x,
                flags=re.I,
            )
        ][0]
        membership_df[membership_column] = (
            membership_df[membership_column]
            .str.replace("\s", " ")
            .str.replace("\xa0", " ", regex=True)
            .str.replace("\u200b", " ", regex=True)
            .apply(lambda x: " ".join(x.split()))
            .str.strip()
        )

        awards_df = pd.concat(awards_list).reset_index(drop=True)
        awards_df = trim_all_columns(awards_df)
        awards_df_obj = awards_df.select_dtypes(["object"])
        awards_df[awards_df_obj.columns] = awards_df_obj.apply(lambda x: x.str.strip())

        license_df = pd.concat(license_list).reset_index(drop=True)
        license_df = trim_all_columns(license_df)
        license_df_obj = license_df.select_dtypes(["object"])
        license_df[license_df_obj.columns] = license_df_obj.apply(lambda x: x.str.strip())
        license_column = [x for x in license_df.columns if re.search("License", x, flags=re.I)][0]
        license_df[license_column] = (
            license_df[license_column]
            .str.replace("\s", " ", regex=True)
            .str.replace("\xa0", " ", regex=True)
            .str.replace("\u200b", " ", regex=True)
            .apply(lambda x: " ".join(x.split()))
            .str.strip()
        )

        # if self.tc_in_simplified_template == True:
        # exp_df.drop(0, inplace=True)
        # degree_df.drop(0, inplace=True)
        # membership_df.drop(0, inplace=True)
        # if len(awards_df) != 0:
        #     awards_df.drop(0, inplace=True)
        # license_df.drop(0, inplace=True)

        a = [
            (exp_df, "Experience"),
            (degree_df, "Degree"),
            (membership_df, "Membership"),
            (awards_df, "Awards"),
            (license_df, "License"),
        ]
        with ExcelWriter(
            self.main_dir + "\\" + "data\\final_processed_data\\{}_details.xlsx".format(self.skg_name),
            mode="w",
            engine="openpyxl",
        ) as writer:
            for df in a:
                df[0].to_excel(writer, sheet_name=df[1], index=False)

    def position_profile_data(self):
        """
        docstring
        """

        # get visible sheets
        wb = load_workbook(filename=self.position_master_data_file_path)

        visible_sheets = []
        for i in wb.worksheets:
            if i.sheet_state == "visible":
                visible_sheets.append(i.title)

        # data_dict = pd.read_excel(self.position_master_data_file_path, sheet_name=None)
        cocode_map_df = pd.read_excel(self.cocode_map_path, sheet_name="Legal Entity Registrations", skiprows=6)

        cocode_map_df.columns = [col.strip() for col in cocode_map_df.columns]
        cocode_map_dict = dict(
            zip(
                cocode_map_df["Legal Entity Identifier"].astype(str).str.zfill(4),
                cocode_map_df["Registered Name"].str.strip(),
            )
        )

        data_dict2 = {}
        for sheet in visible_sheets:
            # if data not in visible_sheets:
            #     # del data_dict[data]
            #     continue

            df = pd.read_excel(self.position_master_data_file_path, sheet_name=sheet)

            if any([re.search("Pos ID", str(x), flags=re.S | re.I) for x in df.columns]):
                pass
            else:
                continue
            df = df[df["Pos ID"].notna()]

            data_dict2[sheet] = df.rename(
                {
                    "Comp. Position": "Company",
                    "Unique Role Id": "SPUR ID",
                    "Unique Role ID": "SPUR ID",
                    "UR ID": "SPUR ID",
                    "UR ID ": "SPUR ID",
                    "Unique Role No.": "SPUR ID",
                    "PID": "Pos ID",
                    "Cocode": "Company ID",
                    "Company Code": "Company ID",
                    "Comp. ID Position": "Company ID",
                    "Comp. ID": "Company ID",
                    "Comp. Position": "Company",
                    "Comp": "Company",
                    "Position Inventory": "Position",
                    "Conso. JG": "Conso JG",
                },
                axis=1,
            )
            data_dict2[sheet].columns = [str(col).strip() for col in data_dict2[sheet]]

        final_data = pd.concat(list(data_dict2.values()))
        final_data = final_data[(final_data["Pos ID"].notna()) & (final_data["SPUR ID"].notna())].reset_index(
            drop=True
        )
        role_level_column = [
            x
            for x in final_data.columns
            if re.search("role level|role level |role\s*level", x.lower().strip(), flags=re.S | re.I)
        ][0]
        if "conso jg" in [x.lower() for x in final_data.columns]:
            conso_jg_column = "Conso JG"
        else:
            conso_jg_column = "JG"

        columns = [
            "Pos ID",
            "SPUR ID",
            "Position",
            role_level_column,
            "Company ID",
            # "Company",
            conso_jg_column,
        ]
        # columns = ['Pos ID', 'SPUR ID', 'Position', 'Role Level', 'Company ID', 'Company', conso_jg_column]
        final_data = final_data[columns].dropna(subset=["Position"])
        final_data[["Pos ID", "Company ID"]] = final_data[["Pos ID", "Company ID"]].astype("str")
        final_data[["Pos ID", "Company ID"]] = final_data[["Pos ID", "Company ID"]].replace(r"\..+", "", regex=True)
        final_data["Pos ID"] = final_data["Pos ID"].str.zfill(8)
        final_data["Company ID"] = final_data["Company ID"].str.zfill(4)
        if final_data["Company ID"].isna().all():
            log.exception("Company ID missing!")
            sys.exit("Please fix the issue and rerun the program")
        if final_data[conso_jg_column].isna().any():
            log.exception("Some JG of the positions are missing!")
            sys.exit("Please fix the issue and rerun the program")
        # invalid_jg_list = []
        # for jg in final_data[conso_jg_column].astype(str):
        #     if jg not in ["A1", "A2","A3","D1","D2", "D3","M1","M2","C1","C2","H1","H2","E3","E4","E5"]:
        #         log.exception("Invalid Job Grade!")
        #     sys.exit("Please fix the issue and rerun the program")
        # if (final_data["Company ID"].isna().all()) or (final_data[conso_jg_column].isna().any()):
        # sys.exit("Please fix the issues and rerun the program")

        final_data = final_data[final_data["Company ID"].notna()]
        final_data["SPUR ID"] = (
            final_data["SPUR ID"].str.replace(" - ", "-", regex=True).str.replace("\s", "", regex=True).str.strip()
        )  # .str.lstrip("0")
        final_data["ProfileCode"] = (
            final_data["SPUR ID"]
            + "_"
            + final_data["Company ID"].astype(str).str.zfill(4)
            + "_"
            + final_data["Pos ID"].astype(str)
        )
        final_data.loc[
            (final_data[role_level_column] == "Custodian") & (final_data[conso_jg_column].isna()),
            conso_jg_column,
        ] = "Custodian"
        # final_data.loc[(final_data['Role Level'] == 'Custodian') & (final_data[conso_jg_column].isna()), conso_jg_column] = 'Custodian'
        final_data[conso_jg_column] = final_data[conso_jg_column].apply(
            lambda x: re.sub(r"^Est.|^Eqv.", r"", str(x)).strip()
        )
        final_data["Company_full_name"] = final_data["Company ID"].astype(str).map(cocode_map_dict)
        final_data.rename(columns={"Conso JG": "JG"}, inplace=True)
        final_data.to_excel(
            self.main_dir + "\\" + "data\\final_processed_data\\{}_position_profile_data.xlsx".format(self.skg_name),
            index=False,
        )

    def tc_data(self):
        """
        docstring
        """
        # final_tc_list = []
        # tc_xlsx_list = glob.glob(self.tc_raw_path)
        # tc_xlsx_list = [x for x in tc_xlsx_list if not re.search(r"~\$", x)]
        # for xlsx in tc_xlsx_list:
        #     data_dict = pd.read_excel(xlsx, sheet_name=None)
        #     for data in list(data_dict.values()):
        #         data_columns = [str(col).strip() for col in data.columns]
        #         if "Oracle" in data_columns or "Compentecy Technical" in data_columns or "ContentItem" in data_columns:
        #             if "A1" in data_columns:
        #                 final_tc_list.append(data)
        #         else:
        #             continue

        # final_tc_data = (
        #     pd.concat(final_tc_list)
        #     .reset_index(drop=True)
        #     .pipe(lambda df: df.loc[:, ~df.columns.str.contains("^Unnamed")])
        #     .dropna(subset="SPUR ID")
        # )
        loc = glob.glob(self.tc_raw_path)[0]
        wb = openpyxl.load_workbook(loc)
        ws = wb.get_sheet_by_name("ProfileItem-Competency TC")

        hidden_cols = []
        for colLetter, colDimension in ws.column_dimensions.items():
            if colDimension.hidden == True:
                hidden_cols.append(string.ascii_lowercase.index(colLetter.lower()))

        df = pd.read_excel(loc, sheet_name="ProfileItem-Competency TC")
        df = df.loc[:, ~df.columns.str.match("Unnamed")]
        unhidden = list(set(df.columns) - set(hidden_cols))
        final_tc_data = df.drop(df.columns[hidden_cols], axis=1).dropna(subset=["SPUR ID"])
        # final_tc_data = df.dropna(subset=["SPUR ID"])
        oracle_column = [
            x
            for x in final_tc_data.columns
            if re.search("ContentItem.1|Oracle|Compentecy Technical|ContentItem", x, flags=re.I)
        ][0]
        final_tc_data[oracle_column] = (
            final_tc_data[oracle_column]
            .apply(lambda x: " ".join(x.split()) if isinstance(x, str) else x)
            .str.replace("–", "-", regex=True)
        )
        numeric_columns = final_tc_data.select_dtypes(include="number").columns
        final_tc_data[numeric_columns] = final_tc_data[numeric_columns].astype("Int64")
        # if self.tc_in_simplified_template == True:
        #     final_tc_data.drop(0, inplace=True)

        final_tc_data["SPUR ID"] = (
            final_tc_data["SPUR ID"].str.replace(" - ", "-", regex=True).str.replace("\s", "", regex=True).str.strip()
        )  # .str.lstrip("0")
        final_tc_data.to_excel(
            self.main_dir + "\\" + "data\\final_processed_data\\{}_TC.xlsx".format(self.skg_name),
            index=False,
        )
